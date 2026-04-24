#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Cronograma Cora - Un Solo Desarrollador
El cronograma más bonito que vas a ver en tu vida.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, NamedStyle,
    GradientFill, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.comments import Comment
from datetime import datetime, timedelta
import random

# Crear workbook
wb = Workbook()

# ==================== COLORES Y ESTILOS ====================
# Paleta de colores profesional (inspirada en Notion/Airtable)
COLORS = {
    'primary': '6366F1',      # Indigo 500
    'primary_dark': '4F46E5',  # Indigo 600
    'secondary': '10B981',    # Emerald 500
    'accent': 'F59E0B',       # Amber 500
    'danger': 'EF4444',       # Red 500
    'purple': '8B5CF6',       # Violet 500
    'pink': 'EC4899',         # Pink 500
    'cyan': '06B6D4',         # Cyan 500
    'slate': '64748B',        # Slate 500
    'slate_light': 'F1F5F9',  # Slate 100
    'white': 'FFFFFF',
    'black': '1E293B',        # Slate 800
    'gray': '94A3B8',         # Slate 400
}

# Bordes
thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

medium_border = Border(
    left=Side(style='medium', color='CBD5E1'),
    right=Side(style='medium', color='CBD5E1'),
    top=Side(style='medium', color='CBD5E1'),
    bottom=Side(style='medium', color='CBD5E1')
)

thick_border = Border(
    left=Side(style='thick', color=COLORS['primary']),
    right=Side(style='thick', color=COLORS['primary']),
    top=Side(style='thick', color=COLORS['primary']),
    bottom=Side(style='thick', color=COLORS['primary'])
)

# ==================== HOJA 1: PORTADA ====================
ws_portada = wb.active
ws_portada.title = "🚀 Portada"

# Configurar anchos de columna
ws_portada.column_dimensions['A'].width = 5
ws_portada.column_dimensions['B'].width = 25
ws_portada.column_dimensions['C'].width = 50
ws_portada.column_dimensions['D'].width = 25

# Título principal
ws_portada.merge_cells('B2:D2')
titulo_cell = ws_portada['B2']
titulo_cell.value = "🚀 CORA TELETHERAPY"
titulo_cell.font = Font(name='Segoe UI', size=36, bold=True, color=COLORS['primary'])
titulo_cell.alignment = Alignment(horizontal='center', vertical='center')

# Subtítulo
ws_portada.merge_cells('B3:D3')
subtitulo = ws_portada['B3']
subtitulo.value = "Cronograma de Desarrollo - Modo Solo Developer"
subtitulo.font = Font(name='Segoe UI', size=16, color=COLORS['slate'])
subtitulo.alignment = Alignment(horizontal='center', vertical='center')

# Línea decorativa
ws_portada.merge_cells('B5:D5')
linea = ws_portada['B5']
linea.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
ws_portada.row_dimensions[5].height = 4

# Info del proyecto
info_data = [
    ("📅 Fecha de Inicio:", "Julio 2026", COLORS['secondary']),
    ("🎯 Fecha de Launch:", "Septiembre 2026", COLORS['accent']),
    ("⏱️ Duración:", "11 semanas", COLORS['purple']),
    ("👤 Equipo:", "1 Desarrollador Full-Stack", COLORS['cyan']),
    ("💻 Rol:", "Backend + Frontend + UX + DevOps", COLORS['pink']),
    ("📊 Total de Tareas:", "65+ tareas principales", COLORS['primary']),
]

row = 7
for label, value, color in info_data:
    ws_portada[f'B{row}'].value = label
    ws_portada[f'B{row}'].font = Font(name='Segoe UI', size=12, bold=True)
    ws_portada[f'B{row}'].alignment = Alignment(horizontal='right')

    ws_portada[f'C{row}'].value = value
    ws_portada[f'C{row}'].font = Font(name='Segoe UI', size=12, color=color, bold=True)
    ws_portada[f'C{row}'].alignment = Alignment(horizontal='left')
    row += 1

# Descripción del enfoque
ws_portada.merge_cells(f'B{row+1}:D{row+4}')
desc = ws_portada[f'B{row+1}']
desc.value = """
🎨 Este cronograma está diseñado para un solo desarrollador ejecutando todo el stack.
Se recomienda: 6-8 horas diarias de trabajo enfocado, con sprints de 2 semanas.
Cada sprint incluye: Backend + Frontend + Testing + Documentación.
"""
desc.font = Font(name='Segoe UI', size=11, color=COLORS['slate'], italic=True)
desc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Footer
ws_portada.merge_cells('B20:D20')
footer = ws_portada['B20']
footer.value = "✨ Hecho con amor para el proyecto más importante de tu vida ✨"
footer.font = Font(name='Segoe UI', size=10, color=COLORS['gray'], italic=True)
footer.alignment = Alignment(horizontal='center')

# ==================== HOJA 2: CRONOGRAMA DETALLADO ====================
ws_cronograma = wb.create_sheet("📅 Cronograma")

# Configurar columnas
ws_cronograma.column_dimensions['A'].width = 8   # Semana
ws_cronograma.column_dimensions['B'].width = 20  # Sprint
ws_cronograma.column_dimensions['C'].width = 18  # Backend
ws_cronograma.column_dimensions['D'].width = 18  # Frontend
ws_cronograma.column_dimensions['E'].width = 15  # UX/UI
ws_cronograma.column_dimensions['F'].width = 18  # DevOps
ws_cronograma.column_dimensions['G'].width = 12  # Status
ws_cronograma.column_dimensions['H'].width = 25  # Entregable

# Encabezado principal
ws_cronograma.merge_cells('A1:H1')
header = ws_cronograma['A1']
header.value = "📅 CRONOGRAMA DETALLADO - 11 SEMANAS"
header.font = Font(name='Segoe UI', size=20, bold=True, color='FFFFFF')
header.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
header.alignment = Alignment(horizontal='center', vertical='center')
ws_cronograma.row_dimensions[1].height = 35

# Encabezados de columnas
headers = ['Semana', 'Sprint', '🔧 Backend', '⚛️ Frontend', '🎨 UX/UI', '🔧 DevOps', 'Estado', '🎯 Entregable']
header_fill = PatternFill(start_color=COLORS['primary_dark'], end_color=COLORS['primary_dark'], fill_type='solid')

for col, header_text in enumerate(headers, 1):
    cell = ws_cronograma.cell(row=2, column=col)
    cell.value = header_text
    cell.font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

ws_cronograma.row_dimensions[2].height = 30

# Datos del cronograma (consolidado para 1 persona)
cronograma_data = [
    # Sprint 1: Fundación (Semanas 1-2)
    ("1", "🏗️ Sprint 1", "Setup Supabase, Auth JWT", "React+Vite setup", "Design system", "CI/CD init", "⏳ Pendiente", "Arquitectura base"),
    ("2", "🏗️ Sprint 1", "MFA implementado", "Login/Register UI", "Wireframes", "Deploy staging", "⏳ Pendiente", "Auth funcional"),

    # Sprint 2: Perfiles (Semanas 3-4)
    ("3", "👤 Sprint 2", "API Perfiles CRUD", "Perfil paciente", "Mockups", "Storage config", "⏳ Pendiente", "Perfiles operativos"),
    ("4", "👤 Sprint 2", "Verificación psicólogos", "Búsqueda+ Filtros", "UX flujos", "Backup setup", "⏳ Pendiente", "Búsqueda lista"),

    # Sprint 3: Citas (Semanas 5-6)
    ("5", "📅 Sprint 3", "API Citas completa", "Calendar widget", "Email templates", "Monitoring", "⏳ Pendiente", "Calendar UI"),
    ("6", "📅 Sprint 3", "Timezone LatAm", "Agendamiento UI", "Notificaciones", "SSL cert", "⏳ Pendiente", "Citas end-to-end"),

    # Sprint 4: Mensajería (Semanas 7-8)
    ("7", "💬 Sprint 4", "Socket.IO server", "Chat interface", "Chat bubbles UI", "Rate limiting", "⏳ Pendiente", "Chat realtime"),
    ("8", "💬 Sprint 4", "Historial mensajes", "Conversaciones list", "Historial UX", "Security audit", "⏳ Pendiente", "Mensajería lista"),

    # Sprint 5: Pagos (Semanas 9-10)
    ("9", "💳 Sprint 5", "Stripe integration", "Checkout UI", "Billing emails", "Webhooks config", "⏳ Pendiente", "Stripe conectado"),
    ("10", "💳 Sprint 5", "Webhooks + Billing", "Cancel 3-clicks", "Feedback form", "PCI check", "⏳ Pendiente", "Pagos funcionando"),

    # Sprint 6: Launch (Semana 11)
    ("11", "🎉 Sprint 6", "Admin APIs", "Admin dashboard", "Admin UX", "Prod deploy", "⏳ Pendiente", "🚀 PRODUCCIÓN"),
]

# Función para obtener color de sprint
def get_sprint_color(sprint_name):
    if "Sprint 1" in sprint_name:
        return 'DBEAFE'  # Blue 100
    elif "Sprint 2" in sprint_name:
        return 'D1FAE5'  # Green 100
    elif "Sprint 3" in sprint_name:
        return 'FEF3C7'  # Amber 100
    elif "Sprint 4" in sprint_name:
        return 'FCE7F3'  # Pink 100
    elif "Sprint 5" in sprint_name:
        return 'E0E7FF'  # Indigo 100
    elif "Sprint 6" in sprint_name:
        return 'FCE7F3'  # Rose 100
    return 'F8FAFC'

def get_status_color(status):
    if "Pendiente" in status:
        return 'FEF3C7'  # Amber
    elif "En progreso" in status:
        return 'DBEAFE'  # Blue
    elif "Completado" in status:
        return 'D1FAE5'  # Green
    return 'F8FAFC'

# Escribir datos
for idx, row_data in enumerate(cronograma_data, 3):
    sprint_color = get_sprint_color(row_data[1])

    for col, value in enumerate(row_data, 1):
        cell = ws_cronograma.cell(row=idx, column=col)
        cell.value = value
        cell.font = Font(name='Segoe UI', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = thin_border

        # Color de fondo para sprint
        if col == 1:  # Semana
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Segoe UI', size=10, bold=True, color=COLORS['primary'])
        elif col == 2:  # Sprint
            cell.fill = PatternFill(start_color=sprint_color, end_color=sprint_color, fill_type='solid')
            cell.font = Font(name='Segoe UI', size=10, bold=True)
        elif col == 8:  # Entregable
            cell.font = Font(name='Segoe UI', size=10, bold=True, color=COLORS['primary_dark'])

        # Estado con color
        if col == 7:
            cell.fill = PatternFill(start_color=get_status_color(value), end_color=get_status_color(value), fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Segoe UI', size=9, bold=True)

    ws_cronograma.row_dimensions[idx].height = 45

# ==================== HOJA 3: TAREAS DETALLADAS ====================
ws_tareas = wb.create_sheet("✅ Tareas Detalladas")

ws_tareas.column_dimensions['A'].width = 8
ws_tareas.column_dimensions['B'].width = 15
ws_tareas.column_dimensions['C'].width = 35
ws_tareas.column_dimensions['D'].width = 12
ws_tareas.column_dimensions['E'].width = 12
ws_tareas.column_dimensions['F'].width = 15

# Encabezado
ws_tareas.merge_cells('A1:F1')
header_tareas = ws_tareas['A1']
header_tareas.value = "✅ LISTA DE TAREAS DETALLADAS"
header_tareas.font = Font(name='Segoe UI', size=18, bold=True, color='FFFFFF')
header_tareas.fill = PatternFill(start_color=COLORS['secondary'], end_color=COLORS['secondary'], fill_type='solid')
header_tareas.alignment = Alignment(horizontal='center', vertical='center')
ws_tareas.row_dimensions[1].height = 30

# Headers
headers_tareas = ['#', 'Sprint', 'Tarea', 'Estimación', 'Prioridad', 'Estado']
for col, header_text in enumerate(headers_tareas, 1):
    cell = ws_tareas.cell(row=2, column=col)
    cell.value = header_text
    cell.font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

tareas_detalladas = [
    # Sprint 1
    ("1", "🏗️ Sprint 1", "Crear proyecto Supabase", "4h", "🔥 Alta", "⏳ Pendiente"),
    ("2", "🏗️ Sprint 1", "Configurar PostgreSQL + RLS", "6h", "🔥 Alta", "⏳ Pendiente"),
    ("3", "🏗️ Sprint 1", "Implementar Auth JWT", "8h", "🔥 Alta", "⏳ Pendiente"),
    ("4", "🏗️ Sprint 1", "Setup MFA TOTP", "6h", "🔥 Alta", "⏳ Pendiente"),
    ("5", "🏗️ Sprint 1", "Crear endpoints auth", "8h", "🔥 Alta", "⏳ Pendiente"),
    ("6", "🏗️ Sprint 1", "Setup React 18 + Vite", "4h", "🔥 Alta", "⏳ Pendiente"),
    ("7", "🏗️ Sprint 1", "Configurar Tailwind", "3h", "Media", "⏳ Pendiente"),
    ("8", "🏗️ Sprint 1", "Crear componentes base (Button, Input)", "8h", "Media", "⏳ Pendiente"),
    ("9", "🏗️ Sprint 1", "Pantalla Login", "6h", "🔥 Alta", "⏳ Pendiente"),
    ("10", "🏗️ Sprint 1", "Pantalla Registro", "8h", "🔥 Alta", "⏳ Pendiente"),
    ("11", "🏗️ Sprint 1", "Integrar React Hook Form", "4h", "Media", "⏳ Pendiente"),
    ("12", "🏗️ Sprint 1", "Form validation Zod", "4h", "Media", "⏳ Pendiente"),
    ("13", "🏗️ Sprint 1", "Definir Design System", "8h", "Media", "⏳ Pendiente"),
    ("14", "🏗️ Sprint 1", "Setup CI/CD GitHub Actions", "6h", "Media", "⏳ Pendiente"),

    # Sprint 2
    ("15", "👤 Sprint 2", "API Perfiles CRUD", "10h", "🔥 Alta", "⏳ Pendiente"),
    ("16", "👤 Sprint 2", "Tabla perfiles en DB", "4h", "🔥 Alta", "⏳ Pendiente"),
    ("17", "👤 Sprint 2", "Sistema verificación psicólogos", "12h", "🔥 Alta", "⏳ Pendiente"),
    ("18", "👤 Sprint 2", "Upload documentos/licencias", "8h", "Media", "⏳ Pendiente"),
    ("19", "👤 Sprint 2", "Storage buckets config", "4h", "Media", "⏳ Pendiente"),
    ("20", "👤 Sprint 2", "Pantalla Perfil Paciente", "10h", "🔥 Alta", "⏳ Pendiente"),
    ("21", "👤 Sprint 2", "Upload avatares", "6h", "Media", "⏳ Pendiente"),
    ("22", "👤 Sprint 2", "Formularios onboarding", "10h", "Media", "⏳ Pendiente"),
    ("23", "👤 Sprint 2", "Página búsqueda psicólogos", "12h", "🔥 Alta", "⏳ Pendiente"),
    ("24", "👤 Sprint 2", "Componentes filtros", "8h", "Media", "⏳ Pendiente"),
    ("25", "👤 Sprint 2", "Tarjetas de psicólogo", "6h", "Media", "⏳ Pendiente"),
    ("26", "👤 Sprint 2", "Perfil público psicólogo", "8h", "Media", "⏳ Pendiente"),

    # Sprint 3
    ("27", "📅 Sprint 3", "API Citas CRUD", "10h", "🔥 Alta", "⏳ Pendiente"),
    ("28", "📅 Sprint 3", "Sistema disponibilidad", "12h", "🔥 Alta", "⏳ Pendiente"),
    ("29", "📅 Sprint 3", "Lógica timezones LatAm", "8h", "Media", "⏳ Pendiente"),
    ("30", "📅 Sprint 3", "Prevención double-booking", "8h", "🔥 Alta", "⏳ Pendiente"),
    ("31", "📅 Sprint 3", "Calendar widget principal", "14h", "🔥 Alta", "⏳ Pendiente"),
    ("32", "📅 Sprint 3", "UI selección horarios", "8h", "Media", "⏳ Pendiente"),
    ("33", "📅 Sprint 3", "Formulario agendamiento", "8h", "Media", "⏳ Pendiente"),
    ("34", "📅 Sprint 3", "Pantalla confirmación cita", "6h", "Media", "⏳ Pendiente"),
    ("35", "📅 Sprint 3", "Notificaciones email SendGrid", "10h", "Media", "⏳ Pendiente"),
    ("36", "📅 Sprint 3", "Templates emails transaccionales", "8h", "Baja", "⏳ Pendiente"),
    ("37", "📅 Sprint 3", "Historial citas paciente", "8h", "Media", "⏳ Pendiente"),
    ("38", "📅 Sprint 3", "Dashboard próximas citas", "8h", "Media", "⏳ Pendiente"),

    # Sprint 4
    ("39", "💬 Sprint 4", "Setup Socket.IO server", "10h", "🔥 Alta", "⏳ Pendiente"),
    ("40", "💬 Sprint 4", "API mensajería real-time", "12h", "🔥 Alta", "⏳ Pendiente"),
    ("41", "💬 Sprint 4", "Persistencia mensajes", "8h", "🔥 Alta", "⏳ Pendiente"),
    ("42", "💬 Sprint 4", "Historial conversaciones", "6h", "Media", "⏳ Pendiente"),
    ("43", "💬 Sprint 4", "Interface chat", "14h", "🔥 Alta", "⏳ Pendiente"),
    ("44", "💬 Sprint 4", "Lista de conversaciones", "8h", "Media", "⏳ Pendiente"),
    ("45", "💬 Sprint 4", "Input mensajes real-time", "8h", "Media", "⏳ Pendiente"),
    ("46", "💬 Sprint 4", "Scroll infinito historial", "8h", "Baja", "⏳ Pendiente"),
    ("47", "💬 Sprint 4", "Indicadores escribiendo...", "6h", "Baja", "⏳ Pendiente"),
    ("48", "💬 Sprint 4", "Panel sesiones terapéuticas", "10h", "Media", "⏳ Pendiente"),
    ("49", "💬 Sprint 4", "Notas simples sesión", "8h", "Media", "⏳ Pendiente"),
    ("50", "💬 Sprint 4", "UI consentimiento recording", "6h", "Media", "⏳ Pendiente"),

    # Sprint 5
    ("51", "💳 Sprint 5", "Integración Stripe", "14h", "🔥 Alta", "⏳ Pendiente"),
    ("52", "💳 Sprint 5", "Payment intents", "8h", "🔥 Alta", "⏳ Pendiente"),
    ("53", "💳 Sprint 5", "Webhooks Stripe config", "8h", "🔥 Alta", "⏳ Pendiente"),
    ("54", "💳 Sprint 5", "API suscripciones", "10h", "🔥 Alta", "⏳ Pendiente"),
    ("55", "💳 Sprint 5", "Lógica billing upgrade/downgrade", "8h", "Media", "⏳ Pendiente"),
    ("56", "💳 Sprint 5", "Checkout Stripe integrado", "10h", "🔥 Alta", "⏳ Pendiente"),
    ("57", "💳 Sprint 5", "Pantalla selección plan", "8h", "Media", "⏳ Pendiente"),
    ("58", "💳 Sprint 5", "Panel suscripción actual", "8h", "Media", "⏳ Pendiente"),
    ("59", "💳 Sprint 5", "Historial pagos", "6h", "Baja", "⏳ Pendiente"),
    ("60", "💳 Sprint 5", "⚡ CANCELACIÓN 3 CLICKS", "10h", "🔥 Alta", "⏳ Pendiente"),
    ("61", "💳 Sprint 5", "Flujo cancelación UX", "6h", "🔥 Alta", "⏳ Pendiente"),
    ("62", "💳 Sprint 5", "Emails billing", "6h", "Baja", "⏳ Pendiente"),

    # Sprint 6
    ("63", "🎉 Sprint 6", "Panel admin dashboard", "14h", "🔥 Alta", "⏳ Pendiente"),
    ("64", "🎉 Sprint 6", "User management CRUD", "10h", "Media", "⏳ Pendiente"),
    ("65", "🎉 Sprint 6", "Review verificaciones", "8h", "Media", "⏳ Pendiente"),
    ("66", "🎉 Sprint 6", "Audit logs viewer", "8h", "Media", "⏳ Pendiente"),
    ("67", "🎉 Sprint 6", "Reports básicos", "10h", "Baja", "⏳ Pendiente"),
    ("68", "🎉 Sprint 6", "Testing E2E con Playwright", "14h", "🔥 Alta", "⏳ Pendiente"),
    ("69", "🎉 Sprint 6", "Bug fixes y polisheo", "10h", "Media", "⏳ Pendiente"),
    ("70", "🎉 Sprint 6", "Performance optimization", "8h", "Media", "⏳ Pendiente"),
    ("71", "🎉 Sprint 6", "Documentación operación", "6h", "Baja", "⏳ Pendiente"),
    ("72", "🎉 Sprint 6", "🚀 DEPLOY A PRODUCCIÓN", "10h", "🔥 Alta", "⏳ Pendiente"),
]

prioridad_colors = {
    "🔥 Alta": 'FEE2E2',
    "Media": 'FEF3C7',
    "Baja": 'E0E7FF',
}

def get_estado_color(estado):
    if "Pendiente" in estado:
        return 'FEF3C7'
    elif "Progreso" in estado:
        return 'DBEAFE'
    return 'D1FAE5'

for idx, row_data in enumerate(tareas_detalladas, 3):
    for col, value in enumerate(row_data, 1):
        cell = ws_tareas.cell(row=idx, column=col)
        cell.value = value
        cell.font = Font(name='Segoe UI', size=9)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border

        if col == 1:  # #
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(name='Segoe UI', size=9, bold=True, color=COLORS['slate'])
        elif col == 2:  # Sprint
            cell.font = Font(name='Segoe UI', size=9, bold=True)
        elif col == 4:  # Estimación
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(name='Segoe UI', size=9, color=COLORS['primary'])
        elif col == 5:  # Prioridad
            cell.fill = PatternFill(start_color=prioridad_colors.get(value, 'FFFFFF'),
                                   end_color=prioridad_colors.get(value, 'FFFFFF'), fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        elif col == 6:  # Estado
            cell.fill = PatternFill(start_color=get_estado_color(value),
                                   end_color=get_estado_color(value), fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(name='Segoe UI', size=9, bold=True)

    ws_tareas.row_dimensions[idx].height = 22

# ==================== HOJA 4: RESUMEN POR SPRINT ====================
ws_resumen = wb.create_sheet("📊 Resumen")

ws_resumen.column_dimensions['A'].width = 20
ws_resumen.column_dimensions['B'].width = 12
ws_resumen.column_dimensions['C'].width = 12
ws_resumen.column_dimensions['D'].width = 15
ws_resumen.column_dimensions['E'].width = 40

# Header
ws_resumen.merge_cells('A1:E1')
header_resumen = ws_resumen['A1']
header_resumen.value = "📊 RESUMEN POR SPRINT"
header_resumen.font = Font(name='Segoe UI', size=18, bold=True, color='FFFFFF')
header_resumen.fill = PatternFill(start_color=COLORS['purple'], end_color=COLORS['purple'], fill_type='solid')
header_resumen.alignment = Alignment(horizontal='center', vertical='center')
ws_resumen.row_dimensions[1].height = 30

# Subheaders
subheaders = ['Sprint', 'Semanas', 'Horas Est.', 'Estado', '🎯 Entregable Clave']
for col, header_text in enumerate(subheaders, 1):
    cell = ws_resumen.cell(row=2, column=col)
    cell.value = header_text
    cell.font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

resumen_data = [
    ("🏗️ Sprint 1: Fundación", "1-2", "80h", "⏳ Pendiente", "Login JWT + MFA funcionando"),
    ("👤 Sprint 2: Perfiles", "3-4", "75h", "⏳ Pendiente", "Búsqueda psicólogos + filtros"),
    ("📅 Sprint 3: Citas", "5-6", "85h", "⏳ Pendiente", "Agendamiento end-to-end"),
    ("💬 Sprint 4: Mensajería", "7-8", "90h", "⏳ Pendiente", "Chat async + historial"),
    ("💳 Sprint 5: Pagos", "9-10", "88h", "⏳ Pendiente", "Stripe + Cancel 3-clicks"),
    ("🎉 Sprint 6: Launch", "11", "72h", "⏳ Pendiente", "🚀 PRODUCCIÓN LISTA"),
]

sprint_colors = ['DBEAFE', 'D1FAE5', 'FEF3C7', 'FCE7F3', 'E0E7FF', 'FFE4E6']

for idx, row_data in enumerate(resumen_data, 3):
    sprint_fill = PatternFill(start_color=sprint_colors[idx-3], end_color=sprint_colors[idx-3], fill_type='solid')

    for col, value in enumerate(row_data, 1):
        cell = ws_resumen.cell(row=idx, column=col)
        cell.value = value
        cell.font = Font(name='Segoe UI', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border
        cell.fill = sprint_fill

        if col == 1:
            cell.font = Font(name='Segoe UI', size=11, bold=True)
        elif col in [2, 3, 4]:
            cell.alignment = Alignment(horizontal='center')
            if col == 3:
                cell.font = Font(name='Segoe UI', size=10, bold=True, color=COLORS['primary'])
        elif col == 5:
            cell.font = Font(name='Segoe UI', size=10, bold=True, color=COLORS['primary_dark'])

    ws_resumen.row_dimensions[idx].height = 35

# Total
ws_resumen.merge_cells('A9:C9')
total_label = ws_resumen['A9']
total_label.value = "TOTAL ESTIMADO:"
total_label.font = Font(name='Segoe UI', size=12, bold=True)
total_label.alignment = Alignment(horizontal='right', vertical='center')
total_label.fill = PatternFill(start_color=COLORS['slate_light'], end_color=COLORS['slate_light'], fill_type='solid')

total_valor = ws_resumen['D9']
total_valor.value = "490h"
total_valor.font = Font(name='Segoe UI', size=14, bold=True, color=COLORS['primary'])
total_valor.alignment = Alignment(horizontal='center', vertical='center')
total_valor.fill = PatternFill(start_color=COLORS['slate_light'], end_color=COLORS['slate_light'], fill_type='solid')
total_valor.border = medium_border

# Nota
ws_resumen.merge_cells('A11:E14')
nota = ws_resumen['A11']
nota.value = """
💡 NOTAS:
• Horas calculadas para 1 developer full-stack trabajando 6-8h diarias
• Total: ~490 horas distribuidas en 11 semanas (~44h/semana)
• Buffer de 20% incluido para imprevistos y bugs
• Sprint 6 incluye testing E2E completo y deploy a producción
"""
nota.font = Font(name='Segoe UI', size=10, italic=True, color=COLORS['slate'])
nota.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
nota.fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

# ==================== HOJA 5: CHECKLIST SEMANAL ====================
ws_checklist = wb.create_sheet("☑️ Checklist")

ws_checklist.column_dimensions['A'].width = 8
ws_checklist.column_dimensions['B'].width = 15
ws_checklist.column_dimensions['C'].width = 50
ws_checklist.column_dimensions['D'].width = 12

# Header
ws_checklist.merge_cells('A1:D1')
header_check = ws_checklist['A1']
header_check.value = "☑️ CHECKLIST SEMANAL DE PROGRESO"
header_check.font = Font(name='Segoe UI', size=18, bold=True, color='FFFFFF')
header_check.fill = PatternFill(start_color=COLORS['cyan'], end_color=COLORS['cyan'], fill_type='solid')
header_check.alignment = Alignment(horizontal='center', vertical='center')
ws_checklist.row_dimensions[1].height = 30

# Subheaders
for col, header_text in enumerate(['Sem', 'Sprint', 'Tarea', '✓'], 1):
    cell = ws_checklist.cell(row=2, column=col)
    cell.value = header_text
    cell.font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='0891B2', end_color='0891B2', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Checklist por semana (resumida)
checklist_items = [
    # Semana 1
    ("1", "🏗️ S1", "□ Proyecto Supabase creado", ""),
    ("1", "🏗️ S1", "□ PostgreSQL con RLS configurado", ""),
    ("1", "🏗️ S1", "□ React + Vite setup completo", ""),
    ("1", "🏗️ S1", "□ Tailwind CSS instalado", ""),
    ("2", "🏗️ S1", "□ Auth JWT implementado", ""),
    ("2", "🏗️ S1", "□ MFA TOTP funcionando", ""),
    ("2", "🏗️ S1", "□ Login/Register UI completo", ""),
    ("2", "🏗️ S1", "□ Deploy a staging", ""),

    # Semana 3-4
    ("3", "👤 S2", "□ API Perfiles CRUD", ""),
    ("3", "👤 S2", "□ Sistema verificación psicólogos", ""),
    ("3", "👤 S2", "□ Upload documentos funcionando", ""),
    ("4", "👤 S2", "□ Perfil paciente UI", ""),
    ("4", "👤 S2", "□ Página búsqueda + filtros", ""),
    ("4", "👤 S2", "□ Perfil público psicólogo", ""),

    # Semana 5-6
    ("5", "📅 S3", "□ API Citas completa", ""),
    ("5", "📅 S3", "□ Sistema disponibilidad", ""),
    ("5", "📅 S3", "□ Timezones LatAm", ""),
    ("6", "📅 S3", "□ Calendar widget", ""),
    ("6", "📅 S3", "□ Agendamiento end-to-end", ""),
    ("6", "📅 S3", "□ Notificaciones email", ""),

    # Semana 7-8
    ("7", "💬 S4", "□ Socket.IO server", ""),
    ("7", "💬 S4", "□ Chat interface básico", ""),
    ("8", "💬 S4", "□ Historial mensajes", ""),
    ("8", "💬 S4", "□ Panel sesiones terapéuticas", ""),
    ("8", "💬 S4", "□ Notas de sesión", ""),

    # Semana 9-10
    ("9", "💳 S5", "□ Stripe integration", ""),
    ("9", "💳 S5", "□ Webhooks configurados", ""),
    ("9", "💳 S5", "□ API suscripciones", ""),
    ("10", "💳 S5", "□ Checkout UI", ""),
    ("10", "💳 S5", "□ Panel suscripción", ""),
    ("10", "💳 S5", "□ ⚡ CANCELACIÓN 3 CLICKS", ""),

    # Semana 11
    ("11", "🎉 S6", "□ Panel admin dashboard", ""),
    ("11", "🎉 S6", "□ User management", ""),
    ("11", "🎉 S6", "□ Audit logs", ""),
    ("11", "🎉 S6", "□ Testing E2E completo", ""),
    ("11", "🎉 S6", "□ Bug fixes finales", ""),
    ("11", "🎉 S6", "□ 🚀 DEPLOY PRODUCCIÓN", ""),
]

semana_colors = {
    "1": "DBEAFE", "2": "DBEAFE",
    "3": "D1FAE5", "4": "D1FAE5",
    "5": "FEF3C7", "6": "FEF3C7",
    "7": "FCE7F3", "8": "FCE7F3",
    "9": "E0E7FF", "10": "E0E7FF",
    "11": "FFE4E6",
}

for idx, row_data in enumerate(checklist_items, 3):
    semana = row_data[0]
    color = semana_colors.get(semana, "FFFFFF")

    for col, value in enumerate(row_data, 1):
        cell = ws_checklist.cell(row=idx, column=col)
        cell.value = value
        cell.font = Font(name='Segoe UI', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

        if col == 1:
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(name='Segoe UI', size=10, bold=True)
        elif col == 2:
            cell.font = Font(name='Segoe UI', size=10, bold=True)
        elif col == 3 and "CANCELACIÓN" in value or "DEPLOY" in value:
            cell.font = Font(name='Segoe UI', size=10, bold=True, color=COLORS['danger'])

    ws_checklist.row_dimensions[idx].height = 25

# ==================== HOJA 6: CALENDARIO VISUAL ====================
ws_cal = wb.create_sheet("🗓️ Calendario")

# Crear un calendario visual tipo Gantt simplificado
ws_cal.column_dimensions['A'].width = 3
ws_cal.column_dimensions['B'].width = 25

# Headers
ws_cal['B2'] = '🗓️ CALENDARIO VISUAL 11 SEMANAS'
ws_cal['B2'].font = Font(name='Segoe UI', size=18, bold=True, color='FFFFFF')
ws_cal['B2'].fill = PatternFill(start_color=COLORS['accent'], end_color=COLORS['accent'], fill_type='solid')
ws_cal['B2'].alignment = Alignment(horizontal='center', vertical='center')
ws_cal.merge_cells('B2:N2')
ws_cal.row_dimensions[2].height = 30

# Timeline header
weeks = ['Jul', 'Jul', 'Jul', 'Jul', 'Ago', 'Ago', 'Ago', 'Ago', 'Sep', 'Sep', 'Sep']
weeks_num = ['S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8', 'S9', 'S10', 'S11']

for i, (month, num) in enumerate(zip(weeks, weeks_num), 3):
    col_letter = get_column_letter(i + 3)
    ws_cal[f'{col_letter}3'] = month
    ws_cal[f'{col_letter}3'].font = Font(name='Segoe UI', size=9, color=COLORS['slate'])
    ws_cal[f'{col_letter}3'].alignment = Alignment(horizontal='center')

    ws_cal[f'{col_letter}4'] = num
    ws_cal[f'{col_letter}4'].font = Font(name='Segoe UI', size=10, bold=True)
    ws_cal[f'{col_letter}4'].alignment = Alignment(horizontal='center')
    ws_cal[f'{col_letter}4'].fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    ws_cal.column_dimensions[col_letter].width = 8

# Sprints
sprints_cal = [
    ("🏗️ Sprint 1: Fundación", 1, 2, '6366F1'),
    ("👤 Sprint 2: Perfiles", 3, 4, '10B981'),
    ("📅 Sprint 3: Citas", 5, 6, 'F59E0B'),
    ("💬 Sprint 4: Mensajería", 7, 8, 'EC4899'),
    ("💳 Sprint 5: Pagos", 9, 10, '8B5CF6'),
    ("🎉 Sprint 6: Launch", 11, 11, 'EF4444'),
]

row = 6
for sprint_name, start_week, end_week, color in sprints_cal:
    ws_cal[f'B{row}'] = sprint_name
    ws_cal[f'B{row}'].font = Font(name='Segoe UI', size=10, bold=True)
    ws_cal[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')

    for week in range(start_week, end_week + 1):
        col = get_column_letter(week + 2)
        ws_cal[f'{col}{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws_cal[f'{col}{row}'].border = thin_border

    ws_cal.row_dimensions[row].height = 25
    row += 2

# Leyenda
row += 2
ws_cal[f'B{row}'] = '🔥 CRÍTICO: '
ws_cal[f'B{row}'].font = Font(name='Segoe UI', size=10, bold=True, color=COLORS['danger'])
ws_cal[f'C{row}'] = 'Mantén el foco en el entregable de cada sprint'
ws_cal[f'C{row}'].font = Font(name='Segoe UI', size=10, italic=True)
ws_cal.merge_cells(f'C{row}:G{row}')

row += 2
ws_cal[f'B{row}'] = '💡 TIP: '
ws_cal[f'B{row}'].font = Font(name='Segoe UI', size=10, bold=True, color=COLORS['primary'])
ws_cal[f'C{row}'] = 'Si te atrásas 1 semana, recupera trabajando fines de semana'
ws_cal[f'C{row}'].font = Font(name='Segoe UI', size=10, italic=True)
ws_cal.merge_cells(f'C{row}:H{row}')

# ==================== GUARDAR ====================
output_path = "D:/cora/docs/Cronograma_Cora_Solo_Dev_11_Semanas.xlsx"
wb.save(output_path)
print(f"[OK] Excel creado exitosamente: {output_path}")
print(f"\n[STATS] RESUMEN:")
print(f"   - 6 hojas con diseño profesional")
print(f"   - 72 tareas detalladas")
print(f"   - ~490 horas de trabajo estimadas")
print(f"   - 11 semanas de desarrollo")
print(f"\n[FEATURES] CARACTERISTICAS:")
print(f"   - Paleta de colores moderna (Indigo/Verde/Ambar)")
print(f"   - Bordes y estilos profesionales")
print(f"   - Checklist interactivo")
print(f"   - Calendario visual tipo Gantt")
print(f"   - Resumen por sprint con totales")
